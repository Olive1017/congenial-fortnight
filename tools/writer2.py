import openpyxl
from openpyxl.utils import get_column_letter

# === 关键字配置 ===
COMPANY_KEY_1 = "P&G"
COMPANY_KEY_2 = "PROCTER & GAMBLE"
HEADER_END_KEYS = ["ITEM NO", "DESCRIPTION"]


def set_smart_print_titles(file_path):
    """
    针对新版 splitter 生成的文件设置打印标题行：
    1. 起始行：包含 P&G 的那一行 (通常是第 1 行)
    2. 结束行：包含 ITEM NO. 的那一行
    """
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        start_row = None
        end_row = None

        # === 1. 扫描行寻找精准边界 ===
        # 扫描前 20 行即可，因为新表头已经压缩了
        for row in ws.iter_rows(min_row=1, max_row=20):
            row_idx = row[0].row
            # 获取当前行所有单元格的文本内容
            row_texts = [str(cell.value).strip().upper() if cell.value else "" for cell in row]
            combined_text = " ".join(row_texts)

            # 寻找起始行：包含 P&G 或公司名
            if start_row is None:
                if COMPANY_KEY_1.upper() in combined_text or COMPANY_KEY_2.upper() in combined_text:
                    start_row = row_idx  # 精准定位，不再 -1

            # 寻找结束行：包含 ITEM NO (注意：判断逻辑要稳健)
            if end_row is None:
                if any("ITEM NO" in t for t in row_texts):
                    end_row = row_idx +1 # 精准定位，不再 +1

            if start_row and end_row:
                break

        # === 2. 执行设置 ===
        if start_row and end_row:
            # 设置打印标题行 (例如 "$1:$5")
            ws.print_title_rows = f"${start_row}:${end_row}"

            # 【额外加固】设置打印区域为整张表（防止打印预览只显示一半）
            # ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

            status_msg = f"成功：固定表头范围 ${start_row}:${end_row}"
            success = True
        else:
            reasons = []
            if not start_row: reasons.append("未找起始行")
            if not end_row: reasons.append("未找结束行(ITEM NO)")
            status_msg = f"失败：{', '.join(reasons)}"
            success = False

        wb.save(file_path)
        wb.close()
        return success, status_msg

    except Exception as e:
        return False, f"发生异常: {str(e)}"