import os
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Font, Alignment, Border, Side

# ========= 配置常量 =========
HEADER_END_KEYS = ["ITEM NO.", "DESCRIPTION"]
COMPANY_KEY_1 = "P&G"
COMPANY_KEY_2 = "PROCTER & GAMBLE (GUANGZHOU) LTD."

# 样式配置
FONT_DELTA = 9
FIRST_ROW_HEIGHT = 36
HEADER_LEFT_COLS = 3  # A-C 列 (P&G 区域)
HEADER_RIGHT_START = 4  # D 列开始
SIGNATURE_BORDER_COLS = (4, 10)  # D 到 I 列
FOOTER_SCAN_ROWS = 15  # 向上扫描的表尾行数

# 页面设置
PAGE_MARGINS_CONFIG = {
    'left': 0, 'right': 0, 'top': 0, 'bottom': 0,
    'header': 0.28, 'footer': 0.12
}

def _row_text(row):
    """提取行文本"""
    return " ".join(str(c.value).strip() for c in row if c.value is not None)

def _is_header_start(row1_text, row2_text):
    """判断是否是表头开始"""
    has_pg = COMPANY_KEY_1 in row1_text
    has_full_name = COMPANY_KEY_2 in row2_text
    return has_pg and has_full_name

def _is_header_end(text):
    """判断是否是表头结束"""
    return all(k in text for k in HEADER_END_KEYS)

def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

# ============== 阶段 1: 解析原表 ============== 
def _detect_tables(ws):
    """
    阶段 1: 检测表格边界
    返回: (header_starts, header_ends, tables, first_table_header_end)
    """
    header_starts = []
    header_ends = []
    rows_list = list(ws.iter_rows(min_row=2))

    # 找表头开始和结束
    for i, row in enumerate(rows_list):
        row_idx = i + 2
        text = _row_text(row)
        if _is_header_end(text):
            header_ends.append(row_idx)
        if i < len(rows_list) - 1:
            if _is_header_start(text, _row_text(rows_list[i + 1])):
                header_starts.append(row_idx)

    if not header_starts or not header_ends:
        raise ValueError("未找到有效表头")

    first_table_header_end = header_ends[0]

    # 修正表头高度（合并单元格检测）
    max_merge_row = first_table_header_end
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= first_table_header_end <= merged.max_row:
            if merged.max_row > max_merge_row:
                max_merge_row = merged.max_row
    first_table_header_end = max_merge_row
    header_ends[0] = max_merge_row

    # 确定表格拆分范围
    tables = []
    for i, header_start in enumerate(header_starts):
        header_end = header_ends[i] if i < len(header_ends) else header_start + 1
        table_start = 1 if i == 0 else header_start
        table_end = header_starts[i + 1] - 2 if i < len(header_starts) - 1 else ws.max_row
        tables.append({
            'start': table_start,
            'end': table_end,
            'header_start': header_start,
            'header_end': header_end
        })

    return header_starts, header_ends, tables, first_table_header_end

# ============== 阶段 2: 准备第一个表的数据 ============== 
def _prepare_first_table_info(ws, tables):
    """
    阶段 2: 准备第一个表的信息
    返回: (first_table_header_row_nums, first_table_last_item_no, first_table_item_col)
    """
    first_table_header_end = tables[0]['header_end']
    
    # 准备表头行号列表
    first_table_header_row_nums = list(range(tables[0]['start'], first_table_header_end + 1))

    # 找 ITEM NO 列
    first_table_item_col = None
    for r_search in range(tables[0]['header_start'], tables[0]['header_end'] + 1):
        for c_idx, cell in enumerate(ws[r_search], 1):
            if cell.value and "ITEM NO" in str(cell.value):
                first_table_item_col = c_idx
                break
        if first_table_item_col:
            break

    # 找最后一个数值
    first_table_last_item_no = None
    if first_table_item_col:
        data_start = tables[0]['header_end'] + 1
        footer_start = None
        for r in range(data_start, tables[0]['end'] + 1):
            if "TOTAL" in _row_text(ws[r]):
                footer_start = r
                break
        end_row = footer_start - 1 if footer_start else tables[0]['end']
        
        for r in range(end_row, data_start - 1, -1):
            try:
                val = ws.cell(r, first_table_item_col).value
                if val is not None:
                    first_table_last_item_no = int(val)
                    break
            except:
                continue

    return first_table_header_row_nums, first_table_last_item_no, first_table_item_col

# ============== 阶段 3: 准备行写入列表 ============== 
def _prepare_rows_to_write(ws, raw_header_rows):
    """
    阶段 3: 准备要写入的行列表（跳过第二行，压缩空白）
    返回: (rows_to_write, original_row2_idx)
    """
    original_row2_idx = None
    if len(raw_header_rows) >= 2:
        original_row2_idx = raw_header_rows[1]
        rows_to_write = [raw_header_rows[0]] + raw_header_rows[2:]
    else:
        rows_to_write = raw_header_rows

    # 压缩空白行逻辑
    def is_blank_old_row(ws_obj, r_idx):
        return all(ws_obj.cell(r_idx, c).value in (None, "") 
                   for c in range(1, ws_obj.max_column + 1))

    compressed_header_rows = []
    prev_blank = False
    for r in rows_to_write:
        if is_blank_old_row(ws, r):
            if not prev_blank:
                compressed_header_rows.append(r)
            prev_blank = True
        else:
            compressed_header_rows.append(r)
            prev_blank = False

    return compressed_header_rows, original_row2_idx

# ============== 阶段 4: 设置页面和样式配置 ============== 
def _setup_page_config(new_ws):
    """设置新工作表的页面配置"""
    margins = PageMargins(**PAGE_MARGINS_CONFIG)
    new_ws.page_margins = margins

    # 打印设置
    new_ws.page_setup.paperSize = new_ws.PAPERSIZE_A4
    new_ws.sheet_properties.pageSetUpPr.fitToPage = True
    new_ws.page_setup.fitToHeight = 0
    new_ws.page_setup.fitToWidth = 1
    new_ws.print_options.horizontalCentered = True

# ============== 阶段 5: 逐行写入 ============== 
def _write_rows(ws, new_ws, rows_to_write, original_row2_idx, max_col):
    """逐行写入数据和样式
    返回: row_map (旧行号 -> 新行号)"""
    def _get_new_col_idx(old_idx):
        if old_idx <= 3:
            return old_idx
        return old_idx + 1

    row_map = {}
    new_r = 1

    for old_r in rows_to_write:
        if old_r in ws.row_dimensions:
            new_ws.row_dimensions[new_r].height = ws.row_dimensions[old_r].height

        is_new_first_row = (new_r == 1)

        for c_idx in range(1, max_col + 1):
            new_c = _get_new_col_idx(c_idx)
            source_cell = ws.cell(old_r, c_idx)

            # 关键修复：第一行 A-C 列智能选择
            if is_new_first_row and new_c <= HEADER_LEFT_COLS:
                if not source_cell.value and original_row2_idx:
                    cell_row2 = ws.cell(original_row2_idx, c_idx)
                    if cell_row2.value:
                        source_cell = cell_row2

            # 写入值和样式
            new_cell = new_ws.cell(row=new_r, column=new_c, value=source_cell.value)
            copy_cell_style(source_cell, new_cell)

            # 复制列宽
            l_old = get_column_letter(c_idx)
            l_new = get_column_letter(new_c)
            if l_old in ws.column_dimensions:
                new_ws.column_dimensions[l_new].width = ws.column_dimensions[l_old].width

        # 补齐 D 列宽度
        if 'C' in ws.column_dimensions:
            new_ws.column_dimensions['D'].width = ws.column_dimensions['C'].width

        row_map[old_r] = new_r
        new_r += 1

    return row_map

# ============== 阶段 6: 处理第一行特殊样式 ============== 
def _apply_first_row_merge_and_style(new_ws, new_r=1):
    """处理第一行：合并左侧、合并右侧、对齐"""
    # 1. 左侧 P&G 合并
    new_ws.merge_cells(start_row=new_r, end_row=new_r, 
                       start_column=1, end_column=HEADER_LEFT_COLS)

    # 2. 右侧长标题合并
    start_col = HEADER_RIGHT_START
    end_col = new_ws.max_column

    # 收集文本
    parts = []
    for c in range(start_col, end_col + 1):
        v = new_ws.cell(new_r, c).value
        if v:
            parts.append(str(v).strip())

    if parts:
        merged_text = "  ".join(parts)
        target_cell = new_ws.cell(new_r, start_col)

        # 找样式源
        style_source_col = start_col
        for c in range(start_col, end_col + 1):
            if new_ws.cell(new_r, c).value:
                style_source_col = c
                break

        # 复制样式并设置
        copy_cell_style(new_ws.cell(new_r, style_source_col), target_cell)
        target_cell.value = merged_text
        target_cell.alignment = Alignment(horizontal="right", vertical="center")

    # 清空右侧多余内容
    for c in range(start_col + 1, end_col + 1):
        new_ws.cell(new_r, c).value = None

    new_ws.merge_cells(start_row=new_r, end_row=new_r, 
                       start_column=start_col, end_column=end_col)

# ============== 阶段 7: 处理合并单元格 ============== 
def _apply_merged_cells(ws, new_ws, row_map):
    """处理原表合并单元格映射到新表"""
    def _get_new_col_idx(old_idx):
        if old_idx <= 3:
            return old_idx
        return old_idx + 1

    for merged in ws.merged_cells.ranges:
        if merged.min_row in row_map and merged.max_row in row_map:
            new_min_c = _get_new_col_idx(merged.min_col)
            new_max_c = _get_new_col_idx(merged.max_col)
            if merged.min_col == 2 and merged.max_col == 3:
                new_max_c = 4

            # 避开第一行
            if not (row_map[merged.min_row] == 1):
                new_ws.merge_cells(
                    start_row=row_map[merged.min_row],
                    end_row=row_map[merged.max_row],
                    start_column=new_min_c,
                    end_column=new_max_c
                )

# ============== 阶段 8: 处理第一行字体放大 ============== 
def _apply_first_row_font_style(new_ws):
    """第一行字体放大和加粗"""
    for c in range(1, new_ws.max_column + 1):
        cell = new_ws.cell(row=1, column=c)
        if cell.value and cell.font:
            new_font = copy(cell.font)
            new_font.size = (new_font.size if new_font.size else 11) + FONT_DELTA
            new_font.bold = True
            cell.font = new_font

    new_ws.row_dimensions[1].height = FIRST_ROW_HEIGHT

# ============== 阶段 9: 找数据起始行 ============== 
def _find_data_start_row(new_ws):
    """找到 ITEM NO 行"""
    for r in range(1, new_ws.max_row + 1):
        row_text = " ".join(str(c.value) for c in new_ws[r] if c.value)
        if "ITEM NO" in row_text:
            return r + 1
    raise ValueError("未找到 ITEM NO 行")

def _find_item_col(new_ws):
    """找到 ITEM NO 列"""
    data_start_row = _find_data_start_row(new_ws)
    for r_h in range(1, data_start_row):
        for c_idx in range(1, new_ws.max_column + 1):
            cell = new_ws.cell(r_h, c_idx)
            if cell.value and "ITEM NO" in str(cell.value):
                return c_idx, data_start_row
    return None, data_start_row

# ============== 阶段 10: 重新编号和处理表尾 ============== 
def _renumber_items_and_format_footer(new_ws, item_col_new, data_start_row):
    """重新编号 ITEM NO 并处理表尾样式"""
    if not item_col_new:
        return

    num = 1
    footer_start_new_row = None
    
    # 找表尾
    for r in range(new_ws.max_row, data_start_row, -1):
        row_txt = "".join([str(c.value or "") for c in new_ws[r]])
        if "TOTAL" in row_txt:
            footer_start_new_row = r
            break

    data_end_row = footer_start_new_row - 1 if footer_start_new_row else new_ws.max_row

    # 重新编号
    for r in range(data_start_row, data_end_row + 1):
        v = new_ws.cell(r, item_col_new).value
        if v is not None:
            new_ws.cell(r, item_col_new).value = num
            num += 1

    # 加粗 TOTAL DAP
    _bold_total_dap_row(new_ws)

def _bold_total_dap_row(new_ws):
    """加粗 TOTAL DAP 及其右侧数值"""
    try:
        found_dap = False
        # 从最后向上扫描
        for r in range(new_ws.max_row, max(1, new_ws.max_row - FOOTER_SCAN_ROWS), -1):
            for c in range(1, new_ws.max_column + 1):
                cell = new_ws.cell(r, c)
                val_str = str(cell.value or "").strip().upper()

                if "TOTAL DAP" in val_str:
                    # 加粗关键字
                    cell.font = Font(
                        bold=True,
                        name=cell.font.name,
                        size=cell.font.size,
                        color=cell.font.color
                    )

                    # 找右侧第一个有值的单元格
                    for target_c in range(c + 1, new_ws.max_column + 1):
                        target_cell = new_ws.cell(r, target_c)
                        if target_cell.value is not None:
                            target_cell.font = Font(
                                bold=True,
                                name=target_cell.font.name,
                                size=target_cell.font.size,
                                color=target_cell.font.color
                            )
                            print(f"   -> 已加粗第 {r} 行的数值单元格: {get_column_letter(target_c)}{r}")
                            break

                    found_dap = True
                    break
            if found_dap:
                break
    except Exception as e:
        print(f"   -> 加粗 TOTAL DAP 时出错: {e}")

# ============== 阶段 11: 添加 signature 边框 ============== 
def _add_signature_border(new_ws):
    """在 SUPPLIER'S ACKNOWLDGMENT 上方加 border"""
    signature_row = None
    
    for r in range(1, new_ws.max_row + 1):
        row_text = " ".join(
            str(new_ws.cell(r, c).value or "").strip().upper()
            for c in range(1, new_ws.max_column + 1)
        )
        # 宽松匹配
        if "SUPPLIER" in row_text and "ACKNOW" in row_text and "MENT" in row_text:
            signature_row = r
            break

    if signature_row:
        thin_black = Side(style="medium", color="000000")
        start_col, end_col = SIGNATURE_BORDER_COLS
        for col in range(start_col, end_col):
            cell = new_ws.cell(signature_row, col)
            current_border = copy(cell.border) if cell.border else Border()
            current_border.top = thin_black
            cell.border = current_border
        print(f"已在第 {signature_row} 行 (SUPPLIER'S ACKNOWLDGMENT) 上方加了 top border")
    else:
        print("警告：仍未找到包含 SUPPLIER & ACKNOW 的行，请检查 debug 输出")

# ============== 主函数 ============== 
def split_excel_by_row(input_path, output_prefix, split_size=30):
    """主流程：拆分 Excel 文件
    split_size 参数保留向后兼容（目前未使用）"""
    # 阶段 1: 解析原表
    wb = load_workbook(input_path)
    ws = wb.active
    max_col = ws.max_column

    header_starts, header_ends, tables, first_table_header_end = _detect_tables(ws)

    # 阶段 2: 准备第一个表的数据
    first_table_header_row_nums, first_table_last_item_no, first_table_item_col = 
        _prepare_first_table_info(ws, tables)

    output_files = []
    output_dir = os.path.dirname(output_prefix)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # 逐个表处理
    for idx, table_info in enumerate(tables, 1):
        new_wb = Workbook()
        new_ws = new_wb.active

        # 阶段 4: 设置页面配置
        _setup_page_config(new_ws)

        # 阶段 3: 准备行列表
        raw_header_rows = list(first_table_header_row_nums)
        rows_to_write, original_row2_idx = _prepare_rows_to_write(ws, raw_header_rows)

        # 确定数据起始行
        if idx == 1:
            data_start_old_row = table_info['header_end'] + 1
        else:
            data_start_old_row = table_info['header_start'] + 1
            if first_table_last_item_no is not None and first_table_item_col:
                exp = first_table_last_item_no + 1
                for r in range(table_info['header_start'], table_info['end'] + 1):
                    try:
                        if ws.cell(r, first_table_item_col).value == exp:
                            data_start_old_row = r
                            break
                    except:
                        continue

        rows_to_write.extend(range(data_start_old_row, table_info['end'] + 1))

        # 阶段 5: 逐行写入
        row_map = _write_rows(ws, new_ws, rows_to_write, original_row2_idx, max_col)

        # 阶段 6: 处理第一行合并
        _apply_first_row_merge_and_style(new_ws)

        # 阶段 7: 处理合并单元格
        _apply_merged_cells(ws, new_ws, row_map)

        # 阶段 8: 处理第一行字体
        _apply_first_row_font_style(new_ws)

        # 阶段 9-10: 重新编号和表尾处理
        item_col_new, data_start_row = _find_item_col(new_ws)
        _renumber_items_and_format_footer(new_ws, item_col_new, data_start_row)

        # 阶段 11: 添加 signature 边框
        _add_signature_border(new_ws)

        # 保存
        suffix = chr(64 + idx)
        if output_prefix.endswith('.xlsx'):
            output_prefix = output_prefix[:-5]
        parts = output_prefix.rsplit(' ', 1)
        out_path = f"{parts[0]}{suffix} {parts[1]}.xlsx" if len(parts) == 2 \
            else f"{output_prefix}{suffix}.xlsx"
        new_wb.save(out_path)
        output_files.append(out_path)
        print(f"✅ {out_path} (样式修复完成)")

    return output_files