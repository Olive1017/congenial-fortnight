import os
import openpyxl
from openpyxl.drawing.image import Image


def add_stamp_to_excel(file_path, stamp_image_path):
    try:
        if not os.path.exists(stamp_image_path):
            return False, f"找不到图片文件: {stamp_image_path}"

        # 使用 with 确保文件安全关闭
        wb = openpyxl.load_workbook(file_path)
        try:
            ws = wb.active
            # 寻找最后一行
            last_row = 1
            for r in range(ws.max_row, 0, -1):
                if any(cell.value is not None for cell in ws[r]):
                    last_row = r
                    break

            img = Image(stamp_image_path)
            img.width, img.height = 180, 126

            # 增加安全边距判断，防止行号为负数
            target_row = max(1, last_row - 4)
            anchor_cell = f"E{target_row}"

            ws.add_image(img, anchor_cell)
            wb.save(file_path)
            return True, f"盖章成功({anchor_cell})"
        finally:
            wb.close()  # 确保在 save 之后或出错后都能关闭

    except Exception as e:
        return False, f"盖章异常: {str(e)}"